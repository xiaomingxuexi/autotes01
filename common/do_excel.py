import openpyxl
from common.constant_path import EXCEL_PATH


class ExcelData:
    def __init__(self,zip_obj):
        for item in zip_obj:
            setattr(self,item[0],item[1])



class DOexcel:

    def __init__(self,file_name,sheet_name):
        self.file_name= file_name
        self.sheet_name= sheet_name
        # self.wb= openpyxl.load_workbook(file_name)
        # self.s= self.wb[self.sheet_name]


    def read_obj(self):# Excel数据加载出来是列表嵌套对象的形式
        # 加载excel
        # 进入sheet页面
        wb = openpyxl.load_workbook(self.file_name)
        s = wb[self.sheet_name]


        # 创建变量
        data_s=[]
        list_head=[]

        # 执行逻辑
        for i in range(1,s.max_row+1):
            # 1.head
            if i ==1:
                for j in range(1,s.max_column+1):
                    value=s.cell(i,j).value
                    if value:
                        list_head.append(value)

            # 2.data
            else:
                # 第i行第一列,是不是每一行的id值:
                if s.cell(i,1).value: # 有值,这一行数据我们就
                    list_data=[]
                    for j in range(1,s.max_column+1):
                        value=s.cell(i,j).value
                        list_data.append(value)
                    ed=ExcelData(zip(list_head,list_data))
                    data_s.append(ed)
        return data_s


    # 写入数据
    def write(self,row,column,value):
        wb = openpyxl.load_workbook(self.file_name)
        s=wb[self.sheet_name]
        s.cell(row,column).value=value
        wb.save(self.file_name)
if __name__ == '__main__':
    sheet="Register"
    do_excel=DOexcel(EXCEL_PATH,sheet)
    dataS=do_excel.read_obj()
    for item in dataS:
        print(item.__dict__)
    # print(dataS)

